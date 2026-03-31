from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

TYPE_LABELS = {
    "cold_rolling": "冷轧",
    "hot_rolling": "热轧",
    "building_steel": "建筑钢材",
    "stainless_flat": "不锈钢平板",
}

LEGACY_HEADERS_NO_PRICE = [
    "类型",
    "一级品类",
    "二级品类",
    "品名",
    "规格",
    "材质",
    "市场",
    "品牌",
    "企业/钢厂",
    "发布时间",
    "日期",
]

BASE_HEADERS = [
    "类型",
    "一级品类",
    "二级品类",
    "品名",
    "规格",
    "材质",
    "市场",
    "品牌",
    "企业/钢厂",
    "价格",
    "发布时间",
    "日期",
]

OUTPUT_HEADERS = ["记录ID", *BASE_HEADERS]

COL_ID = "记录ID"
COL_DATE = "日期"
COL_PRODUCT = "品名"
COL_PRODUCT_KIND = "品种"
COL_SPEC = "规格"
COL_MATERIAL = "材质"
COL_MARKET = "市场"
COL_BRAND = "品牌"
COL_ENTERPRISE = "企业"
COL_MILL = "钢厂"
COL_PRICE = "价格"
UNIT_LABEL = "单位"
DAILY_SUFFIX = "（日）"
SEP = "："
PRICE_COL_INDEX = OUTPUT_HEADERS.index(COL_PRICE) + 1
DATE_COL_INDEX = OUTPUT_HEADERS.index(COL_DATE) + 1
DATE_NUMBER_FORMAT = "yyyy-mm-dd"


@dataclass
class RunMeta:
    execution_strategy: str
    category: str
    subcategory: str
    second_nav: str
    third_nav: str
    product_names: list[str]
    specifications: list[str]
    materials: list[str]
    markets: list[str]
    mills: list[str]
    brands: list[str]
    publish_time: str
    downloaded_file: Path


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def run_date_default() -> str:
    return date.today().isoformat()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_price(value: Any) -> float | None:
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("（均价）", "").replace(",", "").strip()
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def normalize_excel_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text or "~" in text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def date_to_text(value: Any) -> str:
    normalized = normalize_excel_date(value)
    return normalized.isoformat() if normalized else normalize_text(value)


def build_record_id(base_row: list[Any]) -> str:
    normalized_parts: list[str] = []
    for idx, value in enumerate(base_row):
        header = BASE_HEADERS[idx]
        if header == COL_PRICE:
            price = normalize_price(value)
            normalized_parts.append("" if price is None else f"{price:.2f}")
        elif header == COL_DATE:
            normalized_parts.append(date_to_text(value))
        else:
            normalized_parts.append(normalize_text(value))
    payload = "|".join(normalized_parts)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:16]


def with_record_id(base_row: list[Any]) -> list[Any]:
    return [build_record_id(base_row), *base_row]


def iter_today_runs(output_dir: Path, run_date: str) -> list[RunMeta]:
    metas: list[RunMeta] = []
    for path in sorted(output_dir.glob("mysteel_export_*.json")):
        payload = load_json(path)
        captured_at = normalize_text(payload.get("captured_at"))
        if not captured_at.startswith(run_date):
            continue
        query = payload.get("query") or {}
        downloaded_file = Path(normalize_text(payload.get("downloaded_file")))
        if not downloaded_file.exists() or downloaded_file.name.startswith("~$"):
            continue
        metas.append(
            RunMeta(
                execution_strategy=normalize_text(query.get("execution_strategy")),
                category=normalize_text(query.get("category")),
                subcategory=normalize_text(query.get("subcategory")),
                second_nav=normalize_text(query.get("second_nav")),
                third_nav=normalize_text(query.get("third_nav")),
                product_names=[normalize_text(item) for item in (query.get("product_names") or []) if normalize_text(item)],
                specifications=[normalize_text(item) for item in (query.get("specifications") or []) if normalize_text(item)],
                materials=[normalize_text(item) for item in (query.get("materials") or []) if normalize_text(item)],
                markets=[normalize_text(item) for item in (query.get("markets") or []) if normalize_text(item)],
                mills=[normalize_text(item) for item in (query.get("mills") or []) if normalize_text(item)],
                brands=[normalize_text(item) for item in (query.get("brands") or []) if normalize_text(item)],
                publish_time=normalize_text(query.get("publish_time")),
                downloaded_file=downloaded_file,
            )
        )
    return metas


def second_category(meta: RunMeta) -> str:
    return meta.third_nav or meta.subcategory or meta.second_nav


def strip_daily_suffix(value: str) -> str:
    return value.removesuffix(DAILY_SUFFIX).strip()


def parse_matrix_product(meta: RunMeta, product_text: str) -> dict[str, str]:
    parts = [strip_daily_suffix(part) for part in product_text.split(SEP) if strip_daily_suffix(part)]
    result = {
        COL_PRODUCT: "",
        COL_SPEC: "",
        COL_MATERIAL: "",
        COL_MARKET: meta.markets[0] if meta.markets else "",
        COL_BRAND: meta.brands[0] if meta.brands else "",
        "企业/钢厂": meta.mills[0] if meta.mills else "",
    }
    if not parts:
        return result

    result[COL_PRODUCT] = parts[0]
    if meta.execution_strategy == "building_steel":
        if len(parts) > 1:
            result[COL_MATERIAL] = parts[1]
        spec = parts[2] if len(parts) > 2 else ""
        mesh = parts[3] if len(parts) > 3 else ""
        result[COL_SPEC] = " / ".join(item for item in [spec, mesh] if item)
        if len(parts) > 5:
            result[COL_MARKET] = parts[5]
    else:
        if len(parts) > 1:
            result[COL_MATERIAL] = parts[1]
        if len(parts) > 2:
            result[COL_SPEC] = parts[2]
        if len(parts) > 4:
            result[COL_MARKET] = parts[4]
        if len(parts) > 5:
            result["企业/钢厂"] = parts[5]
        if len(parts) > 6:
            result[COL_BRAND] = parts[6]
    return result


def rows_from_matrix_sheet(meta: RunMeta, ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row_idx in range(5, ws.max_row + 1):
        date_value = normalize_excel_date(ws.cell(row_idx, 1).value)
        if date_value is None:
            continue
        for col in range(2, ws.max_column + 1):
            product_text = normalize_text(ws.cell(1, col).value)
            if not product_text:
                continue
            price_value = normalize_price(ws.cell(row_idx, col).value)
            if price_value is None:
                continue
            publish_time = normalize_text(ws.cell(3, col).value) or meta.publish_time
            parsed = parse_matrix_product(meta, product_text)
            base_row = [
                TYPE_LABELS.get(meta.execution_strategy, meta.execution_strategy),
                meta.category,
                second_category(meta),
                parsed[COL_PRODUCT],
                parsed[COL_SPEC],
                parsed[COL_MATERIAL],
                parsed[COL_MARKET],
                parsed[COL_BRAND],
                parsed["企业/钢厂"],
                price_value,
                publish_time,
                date_value,
            ]
            rows.append(with_record_id(base_row))
    return rows


def rows_from_row_sheet(meta: RunMeta, ws) -> list[list[Any]]:
    rows: list[list[Any]] = []
    headers = [normalize_text(ws.cell(4, col).value) for col in range(1, ws.max_column + 1)]
    header_map = {header: idx + 1 for idx, header in enumerate(headers) if header}

    def cell(row_idx: int, name: str) -> str:
        idx = header_map.get(name)
        return normalize_text(ws.cell(row_idx, idx).value) if idx else ""

    publish_time = meta.publish_time or normalize_text(ws.cell(3, 2).value)
    for row_idx in range(5, ws.max_row + 1):
        date_value = normalize_excel_date(cell(row_idx, COL_DATE))
        if date_value is None:
            continue
        product = cell(row_idx, COL_PRODUCT) or cell(row_idx, COL_PRODUCT_KIND) or (meta.product_names[0] if meta.product_names else "")
        spec = cell(row_idx, COL_SPEC) or (meta.specifications[0] if meta.specifications else "")
        material = cell(row_idx, COL_MATERIAL) or (meta.materials[0] if meta.materials else "")
        market = cell(row_idx, COL_MARKET) or (meta.markets[0] if meta.markets else "")
        brand = cell(row_idx, COL_BRAND) or (meta.brands[0] if meta.brands else "")
        mill = cell(row_idx, COL_ENTERPRISE) or cell(row_idx, COL_MILL) or (meta.mills[0] if meta.mills else "")
        price = normalize_price(cell(row_idx, COL_PRICE))
        base_row = [
            TYPE_LABELS.get(meta.execution_strategy, meta.execution_strategy),
            meta.category,
            second_category(meta),
            product,
            spec,
            material,
            market,
            brand,
            mill,
            price,
            publish_time,
            date_value,
        ]
        rows.append(with_record_id(base_row))
    return rows


def extract_rows(meta: RunMeta) -> list[list[Any]]:
    wb = load_workbook(meta.downloaded_file, data_only=True)
    ws = wb.active
    matrix_layout = normalize_text(ws.cell(4, 1).value) == UNIT_LABEL
    return rows_from_matrix_sheet(meta, ws) if matrix_layout else rows_from_row_sheet(meta, ws)


def row_key(row: list[Any]) -> str:
    return normalize_text(row[0])


def upgrade_existing_row(row: list[Any], headers: list[str]) -> list[Any]:
    if headers == OUTPUT_HEADERS:
        upgraded = list(row)
        upgraded[0] = normalize_text(upgraded[0]) or build_record_id(upgraded[1:])
        return upgraded
    if headers == BASE_HEADERS:
        return with_record_id(list(row))
    if headers == LEGACY_HEADERS_NO_PRICE:
        legacy = list(row)
        base_row = [
            legacy[0],
            legacy[1],
            legacy[2],
            legacy[3],
            legacy[4],
            legacy[5],
            legacy[6],
            legacy[7],
            legacy[8],
            None,
            legacy[9],
            legacy[10],
        ]
        return with_record_id(base_row)
    raise RuntimeError("Existing workbook headers do not match expected schema")


def load_existing_rows(output_path: Path) -> list[list[Any]]:
    if not output_path.exists():
        return []
    wb = load_workbook(output_path, data_only=True)
    ws = wb.active
    headers = [normalize_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    supported_headers = (OUTPUT_HEADERS, BASE_HEADERS, LEGACY_HEADERS_NO_PRICE)
    if headers not in supported_headers:
        raise RuntimeError(f"Existing workbook headers do not match expected schema: {output_path}")
    rows: list[list[Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, len(headers) + 1)]
        rows.append(upgrade_existing_row(row, headers))
    return rows


def write_total_price(rows: list[list[Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Total_Price"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, PRICE_COL_INDEX).number_format = "0.00"
        ws.cell(row_idx, DATE_COL_INDEX).number_format = DATE_NUMBER_FORMAT
    ws.freeze_panes = "A2"
    wb.save(output_path)


def merge_rows(existing_rows: list[list[Any]], new_rows: list[list[Any]]) -> tuple[list[list[Any]], int]:
    merged = list(existing_rows)
    existing_keys = {row_key(row) for row in existing_rows}
    added = 0
    for row in new_rows:
        key = row_key(row)
        if key in existing_keys:
            continue
        merged.append(row)
        existing_keys.add(key)
        added += 1
    return merged, added


def main() -> int:
    parser = argparse.ArgumentParser(description="Build Total_Price.xlsx from today's exported Mysteel files")
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--data-dir", default="data")
    parser.add_argument("--run-date", default=run_date_default(), help="Run date in YYYY-MM-DD, default is today")
    parser.add_argument("--output-file", default="Total_Price.xlsx")
    args = parser.parse_args()

    metas = iter_today_runs(Path(args.output_dir), args.run_date)
    if not metas:
        raise RuntimeError(f"No export result json files found for run date: {args.run_date}")

    new_rows: list[list[Any]] = []
    for meta in metas:
        new_rows.extend(extract_rows(meta))
    output_path = Path(args.data_dir) / args.output_file
    existing_rows = load_existing_rows(output_path)
    merged_rows, added_count = merge_rows(existing_rows, new_rows)
    write_total_price(merged_rows, output_path)
    print(f"Built summary workbook: {output_path}")
    print(f"Existing rows: {len(existing_rows)}")
    print(f"Added rows: {added_count}")
    print(f"Total rows: {len(merged_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
